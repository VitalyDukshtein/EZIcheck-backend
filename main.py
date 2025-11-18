from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import tempfile
import os
import re
from collections import Counter
from langdetect import detect, LangDetectException
import langid
from lingua import Language, LanguageDetectorBuilder
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from googletrans import Translator as GoogleTranslator
from html.parser import HTMLParser
import shutil
from pathlib import Path

# Initialize FastAPI app
app = FastAPI(title="EZIcheck-backend API", version="2.0.0")

# Configure CORS (allows your frontend to communicate with backend)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============= Optimized Translation Validation Logic =============

# Precompiled regex patterns for better performance
numeric_pattern = re.compile(r"^[\$€£]?\s*\d+(?:[.,]\d+)?\s*$")
hashtag_phrase_pattern = re.compile(r"#(.*?)#")
hashtag_word_pattern = re.compile(r"#\w+#")

# Precompiled emoji regex pattern
emoji_pattern = re.compile("[" 
    u"\U0001F600-\U0001F64F"  # emoticons
    u"\U0001F300-\U0001F5FF"  # symbols & pictographs
    u"\U0001F680-\U0001F6FF"  # transport & map symbols
    u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
    "]+", flags=re.UNICODE)

# Set of untranslatable terms (using set for O(1) lookup)
UNTRANSLATABLE_TERMS = {"Plus500", "+Premium", "Elite", "Ambassador"}
UNTRANSLATABLE_TERMS_LOWER = {term.lower() for term in UNTRANSLATABLE_TERMS}

# Set of common self-closing HTML tags
SELF_CLOSING_TAGS = {"area", "base", "br", "col", "command", "embed", "hr", "img", 
                     "input", "keygen", "link", "meta", "param", "source", "track", "wbr"}

def extract_emojis(text):
    """Extract emojis from text efficiently"""
    if not isinstance(text, str):
        text = str(text)
    return set(emoji_pattern.findall(text))

# HTMLValidator using Python's HTMLParser
class HTMLValidator(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tag_stack = []
        self.error = False
    
    def handle_starttag(self, tag, attrs):
        if tag not in SELF_CLOSING_TAGS:
            self.tag_stack.append(tag)
    
    def handle_endtag(self, tag):
        if tag in SELF_CLOSING_TAGS:
            return
        if not self.tag_stack or self.tag_stack[-1] != tag:
            self.error = True
        else:
            self.tag_stack.pop()
    
    def validate(self, text):
        self.tag_stack = []
        self.error = False
        try:
            self.feed(text)
            self.close()
        except Exception:
            return False
        return not self.error and not self.tag_stack

# Create a single instance of HTMLValidator to reuse
html_validator = HTMLValidator()

def is_numeric(text):
    """Check if text represents only numbers/amounts"""
    if not isinstance(text, str):
        text = str(text)
    return bool(numeric_pattern.match(text.strip()))

def is_all_hashtag_phrases(text):
    """Check if all tokens in the text are hashtag-wrapped"""
    tokens = text.split()
    return bool(tokens) and all(token.startswith("#") and token.endswith("#") for token in tokens)

def contains_only_untranslatable_terms(text):
    """
    Check if the text contains only terms that shouldn't be translated
    (Plus500, +Premium, Elite, Ambassador, hashtag-wrapped words, or combinations thereof)
    """
    if not isinstance(text, str):
        text = str(text)
    
    words = text.split()
    if not words:
        return False
    
    for word in words:
        word_lower = word.lower()
        # Check if word is an untranslatable term or hashtag-wrapped
        if word_lower not in UNTRANSLATABLE_TERMS_LOWER and not (word.startswith("#") and word.endswith("#")):
            return False
    
    # All words are untranslatable terms
    return True

def check_untranslatable_terms(source_text, target_text):
    """
    Check if terms that should remain unchanged were incorrectly translated.
    Returns (has_error, error_messages)
    """
    if not isinstance(source_text, str) or not isinstance(target_text, str):
        source_text = str(source_text) if source_text is not None else ""
        target_text = str(target_text) if target_text is not None else ""
    
    errors = []
    source_lower = source_text.lower()
    target_lower = target_text.lower()
    
    # Check each untranslatable term
    for term in UNTRANSLATABLE_TERMS:
        term_lower = term.lower()
        if term_lower in source_lower:
            # Count occurrences (case-insensitive)
            source_count = source_lower.count(term_lower)
            target_count = target_lower.count(term_lower)
            
            if target_count < source_count:
                errors.append(f"'{term}' was incorrectly translated or removed")
    
    # Check for hashtag-wrapped words that should remain unchanged
    source_hashtag_words = hashtag_word_pattern.findall(source_text)
    if source_hashtag_words:
        # Convert to set for O(1) lookup
        target_hashtag_set = set(hashtag_word_pattern.findall(target_text))
        for hashtag_word in source_hashtag_words:
            if hashtag_word not in target_hashtag_set:
                errors.append(f"Hashtag-wrapped word '{hashtag_word}' was incorrectly translated or modified")
    
    return len(errors) > 0, errors

def clean_text(text):
    """Clean text by removing untranslatable terms and hashtag-wrapped words"""
    if not isinstance(text, str):
        text = str(text)
    # More efficient: single pass through words
    return [w for w in text.split() 
            if w not in UNTRANSLATABLE_TERMS and not (w.startswith("#") and w.endswith("#"))]

def add_comment_to_cell(cell, comment_text):
    """Add or append a comment to a cell"""
    if cell.comment:
        existing_text = cell.comment.text
        cell.comment = Comment(f"{existing_text}\n{comment_text}", "Translation Validator")
    else:
        cell.comment = Comment(comment_text, "Translation Validator")
    cell.comment.width = 300
    cell.comment.height = 100

# Set up detectors - do this once at module level
supported_languages = list(Language.all())
lingua_detector = LanguageDetectorBuilder.from_languages(*supported_languages).build()
google_translator = GoogleTranslator()

def detect_language(text):
    """Detect language using four libraries"""
    results = []
    
    # Use a single try-except for each detector
    try:
        results.append(detect(text))
    except:
        results.append(None)
    
    try:
        results.append(langid.classify(text)[0])
    except:
        results.append(None)
    
    try:
        lingua_result = lingua_detector.detect_language_of(text)
        results.append(lingua_result.iso_code_639_1.lower() if lingua_result else None)
    except:
        results.append(None)
    
    try:
        detection = google_translator.detect(text)
        results.append(detection.lang.lower())
    except:
        results.append(None)
    
    return results

def detect_language_lingua_only(text):
    """Detect language using only Lingua detector for short texts"""
    try:
        lingua_result = lingua_detector.detect_language_of(text)
        return lingua_result.iso_code_639_1.lower() if lingua_result else None
    except:
        return None

def process_excel_file(file_path: str, output_path: str) -> dict:
    """Process the Excel file and return statistics."""
    
    # Caches to avoid duplicate heavy calls
    detection_cache = {}
    html_cache = {}
    lingua_only_cache = {}
    
    # Statistics
    stats = {
        "total_cells_checked": 0,
        "errors_found": 0,
        "warnings_found": 0,
        "sheets_processed": 0,
        "error_types": {
            "emoji": 0,
            "html": 0,
            "hashtag": 0,
            "language": 0,
            "untranslated": 0,
            "untranslatable": 0,
            "type_mismatch": 0
        }
    }
    
    # Load workbook
    wb = load_workbook(file_path)
    
    # Define cell fills once
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Process each worksheet except "Parameters"
    for sheet in wb.worksheets:
        if sheet.title.strip().lower() == "parameters":
            continue
        
        stats["sheets_processed"] += 1
        max_col = sheet.max_column
        max_row = sheet.max_row
        
        # Process all columns starting from column 3 (including Arabic)
        for col_idx in range(3, max_col + 1):
            header_value = sheet.cell(row=1, column=col_idx).value
            if not header_value:
                continue
            
            lang_code = str(header_value).strip().lower()
            
            # Map language codes to expected values
            if lang_code == "cn":
                expected_lang = "zh_hans"
            elif lang_code == "zh":
                expected_lang = "zh_hant"
            elif lang_code == "nb_no":
                expected_lang = "no"
            elif lang_code == "ar":
                expected_lang = "ar"
            else:
                expected_lang = lang_code
            
            # Process all rows for this column
            for row_idx in range(2, max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                target_text = cell.value
                source_text = sheet.cell(row=row_idx, column=2).value
                
                # Early skip for None values
                if target_text is None and source_text is None:
                    continue
                
                target_str = str(target_text).strip() if target_text is not None else ""
                source_str = str(source_text).strip() if source_text is not None else ""
                
                stats["total_cells_checked"] += 1
                
                # Clear any existing comments
                cell.comment = None
                
                # Skip if target equals ignore phrase (case-insensitive)
                if target_str.lower() == "#cfd_service#. #short_rw#":
                    continue
                
                # Skip if both are numeric and match
                if target_text is not None and source_text is not None:
                    target_is_numeric = isinstance(target_text, (int, float)) or (isinstance(target_text, str) and is_numeric(target_text))
                    source_is_numeric = isinstance(source_text, (int, float)) or (isinstance(source_text, str) and is_numeric(source_text))
                    
                    if target_is_numeric and source_is_numeric and target_str.lower() == source_str.lower():
                        continue
                    
                    # Check type mismatch early
                    if target_is_numeric and not source_is_numeric:
                        cell.fill = red_fill
                        add_comment_to_cell(cell, f"TYPE MISMATCH: Source contains text but target is numeric value '{target_str}'")
                        stats["errors_found"] += 1
                        stats["error_types"]["type_mismatch"] += 1
                        continue
                
                # Skip if target consists solely of hashtag phrases that match source
                if target_str and is_all_hashtag_phrases(target_str) and target_str.lower() == source_str.lower():
                    continue
                
                # Check for incorrectly translated untranslatable terms
                has_untranslatable_error, untranslatable_errors = check_untranslatable_terms(source_str, target_str)
                if has_untranslatable_error:
                    cell.fill = red_fill
                    add_comment_to_cell(cell, "UNTRANSLATABLE TERM ERROR: " + "; ".join(untranslatable_errors))
                    stats["errors_found"] += 1
                    stats["error_types"]["untranslatable"] += 1
                    continue
                
                # Emoji validation (only if source has emojis)
                if source_str and emoji_pattern.search(source_str):
                    source_emojis = extract_emojis(source_str)
                    target_emojis = extract_emojis(target_str)
                    if not source_emojis.issubset(target_emojis):
                        cell.fill = red_fill
                        missing_emojis = source_emojis - target_emojis
                        add_comment_to_cell(cell, f"EMOJI ERROR: Missing emojis from source: {', '.join(missing_emojis)}")
                        stats["errors_found"] += 1
                        stats["error_types"]["emoji"] += 1
                        continue
                
                # HTML validation (only if HTML tags present)
                if "<" in target_str and ">" in target_str:
                    if target_str in html_cache:
                        valid_html = html_cache[target_str]
                    else:
                        valid_html = html_validator.validate(target_str)
                        html_cache[target_str] = valid_html
                    
                    if not valid_html:
                        cell.fill = red_fill
                        add_comment_to_cell(cell, "HTML ERROR: Invalid or unbalanced HTML tags detected")
                        stats["errors_found"] += 1
                        stats["error_types"]["html"] += 1
                        continue
                
                # Hashtag phrases validation (only if source has hashtag phrases)
                if "#" in source_str:
                    source_hashtags = hashtag_phrase_pattern.findall(source_str)
                    if source_hashtags:
                        target_hashtags = hashtag_phrase_pattern.findall(target_str)
                        source_counter = Counter(source_hashtags)
                        target_counter = Counter(target_hashtags)
                        
                        if source_counter != target_counter:
                            cell.fill = red_fill
                            stats["errors_found"] += 1
                            stats["error_types"]["hashtag"] += 1
                            
                            # Build error message efficiently
                            missing_from_target = []
                            extra_in_target = []
                            count_mismatches = []
                            
                            all_hashtags = set(source_hashtags) | set(target_hashtags)
                            for hashtag in all_hashtags:
                                source_count = source_counter.get(hashtag, 0)
                                target_count = target_counter.get(hashtag, 0)
                                
                                if source_count > target_count:
                                    if target_count == 0:
                                        missing_from_target.append(f"#{hashtag}#")
                                    else:
                                        count_mismatches.append(f"#{hashtag}# (expected {source_count}, found {target_count})")
                                elif target_count > source_count:
                                    if source_count == 0:
                                        extra_in_target.append(f"#{hashtag}#")
                                    else:
                                        count_mismatches.append(f"#{hashtag}# (expected {source_count}, found {target_count})")
                            
                            error_messages = []
                            if missing_from_target:
                                error_messages.append(f"Missing: {', '.join(missing_from_target)}")
                            if extra_in_target:
                                error_messages.append(f"Extra: {', '.join(extra_in_target)}")
                            if count_mismatches:
                                error_messages.append(f"Count mismatch: {', '.join(count_mismatches)}")
                            
                            add_comment_to_cell(cell, f"HASHTAG ERROR: {'; '.join(error_messages)}")
                            continue
                
                # Check for untranslated content
                if not target_str or target_str.lower() == source_str.lower():
                    if source_str and not contains_only_untranslatable_terms(source_str):
                        cell.fill = grey_fill
                        stats["warnings_found"] += 1
                        stats["error_types"]["untranslated"] += 1
                        if not target_str:
                            add_comment_to_cell(cell, "UNTRANSLATED: Empty translation")
                        else:
                            add_comment_to_cell(cell, "UNTRANSLATED: Target text matches source text")
                    continue
                
                # Language detection
                cleaned_words = clean_text(target_str)
                if not cleaned_words:  # Skip if no words left after cleaning
                    continue
                
                combined_text = " ".join(cleaned_words)
                word_count = len(cleaned_words)
                
                # Use appropriate detection method based on word count
                if word_count < 4:
                    # For short texts, use only Lingua
                    if combined_text in lingua_only_cache:
                        detected_language = lingua_only_cache[combined_text]
                    else:
                        detected_language = detect_language_lingua_only(combined_text)
                        lingua_only_cache[combined_text] = detected_language
                    
                    if detected_language is not None:
                        # Check for Chinese variants
                        if expected_lang in ("zh_hans", "zh_hant"):
                            match = detected_language == "zh"
                        else:
                            match = detected_language == expected_lang
                        
                        if not match:
                            cell.fill = red_fill
                            stats["errors_found"] += 1
                            stats["error_types"]["language"] += 1
                            add_comment_to_cell(cell, 
                                f"LANGUAGE ERROR (short text): Expected '{expected_lang}', "
                                f"but Lingua detected '{detected_language}' "
                                f"(text has {word_count} word(s) after cleaning)")
                else:
                    # For longer texts, use all 4 detectors
                    if combined_text in detection_cache:
                        detected_languages = detection_cache[combined_text]
                    else:
                        detected_languages = detect_language(combined_text)
                        detection_cache[combined_text] = detected_languages
                    
                    # Check for matches
                    if expected_lang in ("zh_hans", "zh_hant"):
                        match = any(lang == "zh" for lang in detected_languages if lang is not None)
                    else:
                        match = any(lang == expected_lang for lang in detected_languages if lang is not None)
                    
                    if not match and all(lang != expected_lang for lang in detected_languages if lang is not None):
                        cell.fill = red_fill
                        stats["errors_found"] += 1
                        stats["error_types"]["language"] += 1
                        
                        # Format detected languages for the comment
                        detector_names = ["langdetect", "langid", "Lingua", "Google Translate"]
                        detection_results = [f"{detector_names[i]}: {lang}" 
                                           for i, lang in enumerate(detected_languages) 
                                           if lang is not None]
                        
                        add_comment_to_cell(cell, 
                            f"LANGUAGE ERROR: Expected '{expected_lang}', "
                            f"but detected: {', '.join(detection_results)}")
    
    # Save the processed file
    wb.save(output_path)
    return stats

# ============= API Endpoints =============

@app.get("/")
async def root():
    return {
        "message": "EZIcheck-backend API is running!",
        "version": "2.0.0",
        "features": [
            "Language validation with 4 detectors",
            "Emoji preservation check",
            "HTML tag validation", 
            "Hashtag phrase consistency",
            "Untranslatable terms protection",
            "Type mismatch detection",
            "Arabic language support"
        ]
    }

@app.get("/health")
async def health_check():
    return {"status": "healthy", "version": "2.0.0"}

@app.post("/validate")
async def validate_excel(file: UploadFile = File(...)):
    """
    Validate an Excel file for translation errors.
    Returns the processed file with error highlights and comments.
    """
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    # Check file size (10MB limit)
    file_size = 0
    chunk_size = 1024 * 1024  # 1MB chunks
    while chunk := await file.read(chunk_size):
        file_size += len(chunk)
        if file_size > 10 * 1024 * 1024:  # 10MB
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB")
    
    # Reset file position
    await file.seek(0)
    
    # Create temporary directory
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Save uploaded file
        input_path = os.path.join(temp_dir, file.filename)
        with open(input_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Process file
        output_filename = f"validated_{file.filename}"
        output_path = os.path.join(temp_dir, output_filename)
        
        stats = process_excel_file(input_path, output_path)
        
        # Return processed file
        return FileResponse(
            path=output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_filename,
            headers={
                "X-Stats": str(stats),  # Include stats in headers
                "Access-Control-Expose-Headers": "X-Stats"  # Allow frontend to read this header
            }
        )
        
    except Exception as e:
        # Clean up on error
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")
    
    finally:
        # Note: Cleanup will happen after file is sent
        # You might want to implement a background task for cleanup
        pass

@app.post("/validate-with-stats")
async def validate_excel_with_stats(file: UploadFile = File(...)):
    """
    Validate an Excel file and return statistics as JSON.
    Useful for displaying results before download.
    """
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    
    # Check file size
    file_size = 0
    chunk_size = 1024 * 1024
    while chunk := await file.read(chunk_size):
        file_size += len(chunk)
        if file_size > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB")
    
    await file.seek(0)
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Save and process file
        input_path = os.path.join(temp_dir, file.filename)
        with open(input_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        output_filename = f"validated_{file.filename}"
        output_path = os.path.join(temp_dir, output_filename)
        
        stats = process_excel_file(input_path, output_path)
        
        # Read the processed file as base64 for download link
        import base64
        with open(output_path, "rb") as f:
            file_content = f.read()
            file_base64 = base64.b64encode(file_content).decode()
        
        return {
            "filename": output_filename,
            "stats": stats,
            "message": "File processed successfully",
            "file_base64": file_base64  # Frontend can create download link from this
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")
    
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.get("/supported-languages")
async def get_supported_languages():
    """Return list of supported languages for validation."""
    return {
        "languages": [
            {"code": "en", "name": "English"},
            {"code": "es", "name": "Spanish"},
            {"code": "fr", "name": "French"},
            {"code": "de", "name": "German"},
            {"code": "it", "name": "Italian"},
            {"code": "pt", "name": "Portuguese"},
            {"code": "ru", "name": "Russian"},
            {"code": "ar", "name": "Arabic"},
            {"code": "zh", "name": "Chinese (Traditional)"},
            {"code": "cn", "name": "Chinese (Simplified)"},
            {"code": "ja", "name": "Japanese"},
            {"code": "ko", "name": "Korean"},
            {"code": "hi", "name": "Hindi"},
            {"code": "he", "name": "Hebrew"},
            {"code": "tr", "name": "Turkish"},
            {"code": "pl", "name": "Polish"},
            {"code": "nl", "name": "Dutch"},
            {"code": "sv", "name": "Swedish"},
            {"code": "no", "name": "Norwegian"},
            {"code": "da", "name": "Danish"},
            {"code": "fi", "name": "Finnish"},
            {"code": "cs", "name": "Czech"},
            {"code": "hu", "name": "Hungarian"},
            {"code": "ro", "name": "Romanian"},
            {"code": "bg", "name": "Bulgarian"},
            {"code": "el", "name": "Greek"},
            {"code": "th", "name": "Thai"},
            {"code": "vi", "name": "Vietnamese"},
            {"code": "id", "name": "Indonesian"},
            {"code": "ms", "name": "Malay"}
        ]
    }
